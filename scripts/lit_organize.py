#!/usr/bin/env python3
"""
Literature Organizer — organizes academic PDFs into a structured research library.

Modes:
  --scan               : Extract metadata from new PDFs, output JSON summary
  --apply-assignments  : Apply user-confirmed metadata (Phase 1)
  --apply-streams      : Apply stream classifications and create folders (Phase 2)
"""

import argparse
import json
import os
import re
import sys
import hashlib
from datetime import datetime
from difflib import SequenceMatcher
from pathlib import Path

import pdfplumber
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------
STATE_FILE = ".lit-organize-state.json"
SPREADSHEET_COLUMNS = [
    "Title", "Author(s)", "Year", "Source", "Type", "Literature Stream(s)",
    "Research Question / Topic", "Methodology", "Key Findings",
    "Data Sources Used", "Theory/Framework", "Geographic Focus", "Time Period",
    "Why Included", "BibTeX Key", "APA Citation", "In-text Citation",
    "File Path", "Date Added", "Notes",
]
WORKING_PAPER_SERIES = [
    "NBER", "IZA", "CEPR", "SSRN", "BREAD", "CESifo", "World Bank",
    "IMF", "WIDER", "ADB", "ECINEQ",
]

# ---------------------------------------------------------------------------
# Metadata extraction
# ---------------------------------------------------------------------------

def extract_metadata_from_pdf(pdf_path: str) -> dict:
    """Extract title, authors, year, source from PDF metadata and text."""
    meta = {"title": "", "authors": "", "year": "", "source": "", "raw_text_first_pages": ""}

    try:
        with pdfplumber.open(pdf_path) as pdf:
            info = pdf.metadata or {}
            if info.get("Title"):
                meta["title"] = info["Title"].strip()
            if info.get("Author"):
                meta["authors"] = info["Author"].strip()
            if info.get("CreationDate"):
                year_match = re.search(r"(19|20)\d{2}", str(info["CreationDate"]))
                if year_match:
                    meta["year"] = year_match.group()

            # Extract text from first 3 pages
            text_pages = []
            for page in pdf.pages[:3]:
                t = page.extract_text()
                if t:
                    text_pages.append(t)
            full_text = "\n".join(text_pages)
            meta["raw_text_first_pages"] = full_text[:4000]

            # Fallback: parse from text if metadata incomplete
            if not meta["title"] and full_text:
                lines = [l.strip() for l in full_text.split("\n") if l.strip()]
                for line in lines[:5]:
                    if len(line) > 10 and not re.match(r"^(vol|volume|issue|page|doi|http|www|\d)", line, re.I):
                        meta["title"] = line
                        break

            if not meta["authors"] and full_text:
                for line in full_text.split("\n")[:15]:
                    line = line.strip()
                    if re.search(r"\b(and|&)\b", line) and len(line) < 200:
                        if not re.search(r"(abstract|journal|volume|doi|university)", line, re.I):
                            meta["authors"] = line
                            break

            if not meta["year"] and full_text:
                years = re.findall(r"\b(19[5-9]\d|20[0-2]\d)\b", full_text[:1500])
                if years:
                    meta["year"] = years[0]

            # Detect source
            if not meta["source"] and full_text:
                text_lower = full_text[:2000].lower()
                for series in WORKING_PAPER_SERIES:
                    if series.lower() in text_lower:
                        wp_pattern = re.search(
                            rf"{re.escape(series)}[^0-9]*(\d+)", full_text[:2000], re.I
                        )
                        if wp_pattern:
                            meta["source"] = f"{series} Working Paper {wp_pattern.group(1)}"
                        else:
                            meta["source"] = f"{series} Working Paper"
                        break

                if not meta["source"]:
                    journal_patterns = [
                        r"(?:published in|forthcoming in|accepted at)\s+(.+?)[\.\,\n]",
                        r"((?:American|Quarterly|Journal|Review|Economic)[^\n,]{5,60})",
                    ]
                    for pat in journal_patterns:
                        m = re.search(pat, full_text[:2000], re.I)
                        if m:
                            meta["source"] = m.group(1).strip()
                            break

    except Exception as e:
        meta["error"] = str(e)

    return meta


def compute_file_hash(path: str) -> str:
    h = hashlib.md5()
    with open(path, "rb") as f:
        for chunk in iter(lambda: f.read(8192), b""):
            h.update(chunk)
    return h.hexdigest()


# ---------------------------------------------------------------------------
# Naming and deduplication
# ---------------------------------------------------------------------------

def sanitize_part(s: str, max_len: int = 50) -> str:
    """Sanitize a single part of the filename. Spaces become spaces, not underscores."""
    s = re.sub(r"[^\w\s-]", "", s)
    s = re.sub(r"\s+", " ", s.strip())
    return s[:max_len]


def sanitize_filename(s: str, max_len: int = 50) -> str:
    """Legacy: sanitize for folder names (spaces to underscores)."""
    s = re.sub(r"[^\w\s-]", "", s)
    s = re.sub(r"\s+", "_", s.strip())
    return s[:max_len]


def make_filename(title: str, author: str, year: str, source: str, number: int = 0) -> str:
    t = sanitize_part(title, 60)
    a = author.split(",")[0].split(" and ")[0].split("&")[0].strip()
    a_last = a.split()[-1] if a.split() else "Unknown"
    a_last = sanitize_part(a_last, 30)
    y = year if year else "NoYear"
    prefix = f"{number}_" if number > 0 else ""
    return f"{prefix}{t}_{a_last}_{y}.pdf"


def title_similarity(t1: str, t2: str) -> float:
    return SequenceMatcher(None, t1.lower().strip(), t2.lower().strip()).ratio()


def is_working_paper(source: str) -> bool:
    if not source:
        return False
    return any(s.lower() in source.lower() for s in WORKING_PAPER_SERIES)


# ---------------------------------------------------------------------------
# BibTeX and citations
# ---------------------------------------------------------------------------

def make_bibtex_key(author: str, year: str, title: str) -> str:
    a = author.split(",")[0].split(" and ")[0].split("&")[0].strip()
    a_last = a.split()[-1].lower() if a.split() else "unknown"
    a_last = re.sub(r"[^a-z]", "", a_last)
    y = year if year else "nd"
    title_words = [w.lower() for w in re.findall(r"[A-Za-z]+", title)]
    stop = {"the", "a", "an", "of", "in", "on", "and", "for", "to", "with", "from"}
    first_word = next((w for w in title_words if w not in stop), "untitled")
    return f"{a_last}{y}{first_word}"


def make_apa_citation(author: str, year: str, title: str, source: str) -> str:
    y = year if year else "n.d."
    s = f" {source}." if source else ""
    return f"{author} ({y}). {title}.{s}"


def make_intext_citation(author: str, year: str) -> str:
    a = author.split(",")[0].split(" and ")[0].split("&")[0].strip()
    a_last = a.split()[-1] if a.split() else "Unknown"
    y = year if year else "n.d."
    return f"({a_last}, {y})"


def make_bibtex_entry(key: str, author: str, year: str, title: str, source: str, entry_type: str) -> str:
    bib_type = "article"
    if "Working Paper" in (entry_type or ""):
        bib_type = "techreport"
    elif "Book Chapter" in (entry_type or ""):
        bib_type = "incollection"
    elif "Book" in (entry_type or ""):
        bib_type = "book"

    lines = [f"@{bib_type}{{{key},"]
    lines.append(f"  author = {{{author}}},")
    lines.append(f"  title = {{{title}}},")
    if year:
        lines.append(f"  year = {{{year}}},")
    if source:
        if bib_type == "article":
            lines.append(f"  journal = {{{source}}},")
        elif bib_type == "techreport":
            lines.append(f"  institution = {{{source}}},")
        else:
            lines.append(f"  publisher = {{{source}}},")
    lines.append("}")
    return "\n".join(lines)


# ---------------------------------------------------------------------------
# State management
# ---------------------------------------------------------------------------

def load_state(folder: str) -> dict:
    path = os.path.join(folder, STATE_FILE)
    if os.path.exists(path):
        with open(path) as f:
            return json.load(f)
    return {"processed_files": {}, "entries": []}


def save_state(folder: str, state: dict):
    path = os.path.join(folder, STATE_FILE)
    with open(path, "w") as f:
        json.dump(state, f, indent=2)


# ---------------------------------------------------------------------------
# Spreadsheet
# ---------------------------------------------------------------------------

def get_or_create_workbook(folder: str, project: str) -> tuple:
    xlsx_path = os.path.join(folder, f"{project}_literature.xlsx")
    if os.path.exists(xlsx_path):
        wb = openpyxl.load_workbook(xlsx_path)
        ws = wb.active
        return wb, ws, xlsx_path

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Literature"

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    for col_idx, col_name in enumerate(SPREADSHEET_COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = thin_border

    widths = {
        "A": 35, "B": 25, "C": 8, "D": 25, "E": 15, "F": 20,
        "G": 30, "H": 20, "I": 35, "J": 20, "K": 20, "L": 15,
        "M": 15, "N": 25, "O": 20, "P": 40, "Q": 18, "R": 30,
        "S": 12, "T": 25,
    }
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = "A1:T1"

    wb.save(xlsx_path)
    return wb, ws, xlsx_path


def append_to_spreadsheet(ws, entry: dict, row: int):
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    values = [
        entry.get("title", ""),
        entry.get("authors", ""),
        entry.get("year", ""),
        entry.get("source", ""),
        entry.get("type", ""),
        entry.get("streams", ""),
        entry.get("research_question", ""),
        entry.get("methodology", ""),
        entry.get("key_findings", ""),
        entry.get("data_sources", ""),
        entry.get("theory", ""),
        entry.get("geographic_focus", ""),
        entry.get("time_period", ""),
        "",  # Why Included — left blank
        entry.get("bibtex_key", ""),
        entry.get("apa_citation", ""),
        entry.get("intext_citation", ""),
        entry.get("file_path", ""),
        entry.get("date_added", ""),
        entry.get("notes", ""),
    ]
    for col_idx, val in enumerate(values, 1):
        cell = ws.cell(row=row, column=col_idx, value=val)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        cell.border = thin_border


# ---------------------------------------------------------------------------
# BibTeX file
# ---------------------------------------------------------------------------

def append_to_bib(folder: str, project: str, bib_entry: str):
    bib_path = os.path.join(folder, f"{project}_literature.bib")
    with open(bib_path, "a") as f:
        f.write("\n" + bib_entry + "\n")


# ---------------------------------------------------------------------------
# Main commands
# ---------------------------------------------------------------------------

def check_deletions(folder: str, project: str):
    """Check for files that were deleted from disk and mark them in the spreadsheet."""
    state = load_state(folder)
    xlsx_path = os.path.join(folder, f"{project}_literature.xlsx")
    if not os.path.exists(xlsx_path):
        return []

    deleted = []
    for entry in state.get("entries", []):
        fp = entry.get("file_path", "")
        if fp and not os.path.exists(os.path.join(folder, fp)):
            deleted.append(fp)

    if not deleted:
        return []

    # Mark deleted in spreadsheet
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active
    for row in range(2, ws.max_row + 1):
        fp = ws.cell(row=row, column=18).value  # File Path column
        if fp in deleted:
            notes = ws.cell(row=row, column=20).value or ""
            if "DELETED" not in notes:
                ws.cell(row=row, column=20, value=f"DELETED. {notes}".strip())

    wb.save(xlsx_path)

    # Remove from state entries and processed_files
    state["entries"] = [e for e in state["entries"] if e.get("file_path") not in deleted]
    state["processed_files"] = {k: v for k, v in state["processed_files"].items() if v not in deleted}
    save_state(folder, state)

    # Remove dangling symlinks in stream folders
    for d in os.listdir(folder):
        dpath = os.path.join(folder, d)
        if os.path.isdir(dpath) and not d.startswith('.'):
            for f in os.listdir(dpath):
                link = os.path.join(dpath, f)
                if os.path.islink(link) and not os.path.exists(link):
                    os.unlink(link)

    return deleted


def cmd_scan(folder: str, project: str):
    """Scan for new PDFs and output metadata as JSON."""
    # Check for deletions first
    deleted = check_deletions(folder, project)

    state = load_state(folder)
    processed = set(state["processed_files"].keys())

    pdfs = sorted(Path(folder).glob("*.pdf"))
    new_pdfs = [p for p in pdfs if p.name not in processed and not p.is_symlink()]

    if not new_pdfs and not deleted:
        print(json.dumps({"new_files": [], "deleted_files": [], "message": "No new PDFs found."}))
        return
    if not new_pdfs:
        print(json.dumps({"new_files": [], "deleted_files": deleted, "message": f"No new PDFs. {len(deleted)} deleted files marked."}))
        return

    results = []
    for pdf_path in new_pdfs:
        meta = extract_metadata_from_pdf(str(pdf_path))
        file_hash = compute_file_hash(str(pdf_path))
        results.append({
            "original_filename": pdf_path.name,
            "file_hash": file_hash,
            "title": meta.get("title", ""),
            "authors": meta.get("authors", ""),
            "year": meta.get("year", ""),
            "source": meta.get("source", ""),
            "raw_text_preview": meta.get("raw_text_first_pages", "")[:800],
        })

    # Check for duplicates
    duplicates = []
    existing_entries = state.get("entries", [])

    for r in results:
        for e in existing_entries:
            if r["file_hash"] == e.get("file_hash"):
                duplicates.append({"new_file": r["original_filename"], "existing": e.get("file_path", ""), "reason": "exact_duplicate"})
                break
            if r["title"] and e.get("title") and title_similarity(r["title"], e["title"]) > 0.85:
                new_is_wp = is_working_paper(r["source"])
                existing_is_wp = is_working_paper(e.get("source", ""))
                if new_is_wp and not existing_is_wp:
                    duplicates.append({"new_file": r["original_filename"], "existing": e.get("file_path", ""), "reason": "new_is_working_paper"})
                elif existing_is_wp and not new_is_wp:
                    duplicates.append({"new_file": r["original_filename"], "existing": e.get("file_path", ""), "reason": "existing_is_working_paper"})
                else:
                    duplicates.append({"new_file": r["original_filename"], "existing": e.get("file_path", ""), "reason": "title_match_duplicate"})
                break

    output = {
        "new_files": results,
        "duplicates": duplicates,
        "deleted_files": deleted,
        "project": project,
        "folder": folder,
    }
    print(json.dumps(output, indent=2))


def cmd_apply(folder: str, project: str, assignments_path: str):
    """Apply confirmed metadata and stream assignments."""
    with open(assignments_path) as f:
        assignments = json.load(f)

    state = load_state(folder)
    wb, ws, xlsx_path = get_or_create_workbook(folder, project)
    next_row = ws.max_row + 1

    # Handle duplicates
    for dup in assignments.get("duplicates_to_delete", []):
        dup_path = os.path.join(folder, dup)
        if os.path.exists(dup_path):
            os.remove(dup_path)
            print(f"Deleted duplicate: {dup}")

    # Handle working paper replacements
    for replacement in assignments.get("working_paper_replacements", []):
        wp_path = os.path.join(folder, replacement["delete"])
        if os.path.exists(wp_path):
            os.remove(wp_path)
            print(f"Deleted working paper (published version exists): {replacement['delete']}")
            for entry in state["entries"]:
                if entry.get("original_filename") == replacement["delete"]:
                    entry["replaced_by"] = replacement.get("keep", "")

    # Determine next numbering: find highest existing number prefix
    existing_numbers = []
    for fname in os.listdir(folder):
        m = re.match(r"^(\d+)_", fname)
        if m:
            existing_numbers.append(int(m.group(1)))
    next_number = max(existing_numbers) + 1 if existing_numbers else 1

    # Process each confirmed entry
    for item in assignments.get("entries", []):
        original = item["original_filename"]
        title = item["title"]
        authors = item["authors"]
        year = item["year"]
        source = item["source"]
        entry_type = item.get("type", "Journal Article")
        streams = item.get("streams", [])
        streams_str = ", ".join(streams)

        new_filename = make_filename(title, authors, year, source, number=next_number)
        next_number += 1
        bibtex_key = make_bibtex_key(authors, year, title)
        apa = make_apa_citation(authors, year, title, source)
        intext = make_intext_citation(authors, year)
        bib_entry = make_bibtex_entry(bibtex_key, authors, year, title, source, entry_type)

        # Rename file
        old_path = os.path.join(folder, original)
        new_path = os.path.join(folder, new_filename)
        if os.path.exists(old_path):
            if os.path.exists(new_path) and old_path != new_path:
                base, ext = os.path.splitext(new_filename)
                counter = 2
                while os.path.exists(os.path.join(folder, f"{base}_{counter}{ext}")):
                    counter += 1
                new_filename = f"{base}_{counter}{ext}"
                new_path = os.path.join(folder, new_filename)
            os.rename(old_path, new_path)
            print(f"Renamed: {original} -> {new_filename}")

        entry = {
            "title": title,
            "authors": authors,
            "year": year,
            "source": source,
            "type": entry_type,
            "streams": streams_str,
            "research_question": item.get("research_question", ""),
            "methodology": item.get("methodology", ""),
            "key_findings": item.get("key_findings", ""),
            "data_sources": item.get("data_sources", ""),
            "theory": item.get("theory", ""),
            "geographic_focus": item.get("geographic_focus", ""),
            "time_period": item.get("time_period", ""),
            "bibtex_key": bibtex_key,
            "apa_citation": apa,
            "intext_citation": intext,
            "file_path": new_filename,
            "date_added": datetime.now().strftime("%Y-%m-%d"),
            "notes": item.get("notes", ""),
        }

        append_to_spreadsheet(ws, entry, next_row)
        next_row += 1
        append_to_bib(folder, project, bib_entry)

        state["processed_files"][original] = new_filename
        file_hash = compute_file_hash(new_path) if os.path.exists(new_path) else ""
        state["entries"].append({
            "original_filename": original,
            "file_path": new_filename,
            "file_hash": file_hash,
            "title": title,
            "authors": authors,
            "year": year,
            "source": source,
            "bibtex_key": bibtex_key,
        })

    wb.save(xlsx_path)
    save_state(folder, state)
    print(f"\nDone! Processed {len(assignments.get('entries', []))} files.")
    print(f"Spreadsheet: {xlsx_path}")
    print(f"BibTeX: {os.path.join(folder, project + '_literature.bib')}")


def cmd_apply_streams(folder: str, project: str, streams_path: str):
    """Apply stream classifications: update spreadsheet column and create symlink folders."""
    with open(streams_path) as f:
        data = json.load(f)

    state = load_state(folder)
    xlsx_path = os.path.join(folder, f"{project}_literature.xlsx")
    if not os.path.exists(xlsx_path):
        print(f"Error: spreadsheet not found at {xlsx_path}", file=sys.stderr)
        sys.exit(1)

    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active

    # Build a map: file_path -> row number (column R = 18 = File Path)
    file_to_row = {}
    for row in range(2, ws.max_row + 1):
        fp = ws.cell(row=row, column=18).value
        if fp:
            file_to_row[fp] = row

    for assignment in data.get("stream_assignments", []):
        file_path = assignment["file_path"]
        streams = assignment.get("streams", [])
        streams_str = ", ".join(streams)

        # Update spreadsheet column F (Literature Stream(s) = column 6)
        row = file_to_row.get(file_path)
        if row:
            ws.cell(row=row, column=6, value=streams_str)
            print(f"Updated streams for {file_path}: {streams_str}")
        else:
            print(f"Warning: {file_path} not found in spreadsheet, skipping column update")

        # Create symlink folders
        for stream in streams:
            stream_dir = os.path.join(folder, sanitize_filename(stream, 80))
            os.makedirs(stream_dir, exist_ok=True)
            link_path = os.path.join(stream_dir, file_path)
            if not os.path.exists(link_path):
                target = os.path.join("..", file_path)
                os.symlink(target, link_path)
                print(f"  Linked {file_path} -> {stream_dir}/")

        # Update state entries
        for entry in state.get("entries", []):
            if entry.get("file_path") == file_path:
                entry["streams"] = streams_str
                break

    wb.save(xlsx_path)
    save_state(folder, state)
    print(f"\nDone! Classified {len(data.get('stream_assignments', []))} papers into streams.")


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(description="Literature Organizer")
    parser.add_argument("--folder", required=True, help="Folder containing PDFs")
    parser.add_argument("--project", default="research", help="Project name")
    parser.add_argument("--scan", action="store_true", help="Scan for new PDFs")
    parser.add_argument("--apply-assignments", help="Path to assignments JSON (Phase 1)")
    parser.add_argument("--apply-streams", help="Path to streams JSON (Phase 2)")

    args = parser.parse_args()

    if not os.path.isdir(args.folder):
        print(f"Error: {args.folder} is not a directory", file=sys.stderr)
        sys.exit(1)

    if args.apply_streams:
        cmd_apply_streams(args.folder, args.project, args.apply_streams)
    elif args.apply_assignments:
        cmd_apply(args.folder, args.project, args.apply_assignments)
    else:
        cmd_scan(args.folder, args.project)


if __name__ == "__main__":
    main()
